import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Function to open a file dialog for selecting an image file
def select_image_file():
    root = Tk()
    root.withdraw()  # Hide the main window
    file_path = askopenfilename(title="Select Image File", filetypes=[("Image files", "*.jpg *.jpeg *.png")])
    return file_path

# Define the findEncodings function
def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList

# Define the markAttendance function
def markAttendance(names):
    # Get the current date
    now = datetime.now()
    date_string = now.strftime('%Y-%m-%d')

    excel_file = f'attendance_{date_string}.xlsx'
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Time"])
        wb.save(excel_file)

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for name in names:
        nameList = []
        for row in ws.iter_rows(values_only=True):
            nameList.append(row[0])

        if name not in nameList:
            now = datetime.now()
            dtString = now.strftime('%H:%M:%S')
            ws.append([name, dtString])
            wb.save(excel_file)

# Function to process the selected image file
def process_image_file(file_path):
    image = cv2.imread(file_path)

    # Load known face encodings from the ImagesAttendance folder
    path = 'ImagesAttendance'
    encodeListKnown = []
    classNames = []

    for cl in os.listdir(path):
        curImg = cv2.imread(os.path.join(path, cl))
        encode = face_recognition.face_encodings(curImg)[0]
        encodeListKnown.append(encode)
        classNames.append(os.path.splitext(cl)[0])

    imgS = cv2.resize(image, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    names = []

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            name = classNames[matchIndex].upper()
            names.append(name)

    if names:
        markAttendance(names)
        print("Attendance updated for:", names)
    else:
        print("No faces recognized in the uploaded image.")

# Call the function to select an image file and process it
if __name__ == "__main__":
    selected_file = select_image_file()
    if selected_file:
        process_image_file(selected_file)
